from matplotlib.table import Cell
import pandas as pd
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment
 
from Asignacion_Ciclo.models import AsignacionCiclo, Promocion
from Curso.models import Curso, Grado
from Persona.models import Alumna
from .models import Actividad, CalificacionActividad
from django.db.models import Sum
from django.db.models import Avg
from datetime import datetime

def exportar_calificaciones_excel(request, alumna_id):
    alumna = get_object_or_404(Alumna, id=alumna_id)
    
    # Obtener los filtros
    grado_id = request.GET.get('grado')
    year = request.GET.get('year')
    curso_id = request.GET.get('curso')

    # Obtener calificaciones según los filtros
    calificaciones = CalificacionActividad.objects.filter(asignacion_ciclo__alumna=alumna).order_by('-actividad__fecha')

    if grado_id:
        calificaciones = calificaciones.filter(asignacion_ciclo__grado_id=grado_id)
    if year:
        calificaciones = calificaciones.filter(asignacion_ciclo__year=year)
    if curso_id:
        calificaciones = calificaciones.filter(actividad__curso_id=curso_id)
    
    # Crear un DataFrame con la información de la alumna
    datos_alumna = {
        "Código": [alumna.codigo],
        "Nombre": [alumna.persona.nombre],
        "Apellido": [alumna.persona.apellido]
    }
    
    df_alumna = pd.DataFrame(datos_alumna)
    
    # Crear un DataFrame con las calificaciones
    datos_calificaciones = []
    for calificacion in calificaciones:
        datos_calificaciones.append({
            "Actividad": calificacion.actividad.actividad,
            "Curso": calificacion.actividad.curso.nombre_curso,
            "Fecha": calificacion.actividad.fecha,
            "Descripción": calificacion.descripcion,
            "Punteo": calificacion.punteo,
            "Punteo Total": calificacion.actividad.punteo,
            "Ciclo": calificacion.asignacion_ciclo.grado.nombre_grado,
            "Año": calificacion.asignacion_ciclo.year
        })
    
    df_calificaciones = pd.DataFrame(datos_calificaciones)

    # Crear un archivo Excel con la información
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="calificaciones_alumna.xlsx"'

    # Crear un libro y una hoja de trabajo
    wb = Workbook()
    ws = wb.active
    ws.title = 'Calificaciones Alumna'

    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="008000", end_color="008000", fill_type="solid")
    center_aligned_text = Alignment(horizontal="center")
    data_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")

    # Escribir los encabezados de la alumna
    headers_alumna = ["Código", "Nombre", "Apellido"]
    for col_num, header in enumerate(headers_alumna, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_aligned_text

    # Escribir los datos de la alumna
    for col_num, (header, value) in enumerate(datos_alumna.items(), 1):
        cell = ws.cell(row=2, column=col_num, value=value[0] if isinstance(value, list) else value)
        cell.fill = data_fill
        cell.alignment = center_aligned_text

    # Escribir una fila en blanco para separar los datos de la alumna de la tabla de calificaciones
    start_row_calificaciones = len(df_alumna) + 3

    # Escribir los encabezados de la tabla de calificaciones
    headers_calificaciones = ["Actividad", "Curso", "Fecha", "Descripción", "Punteo", "Punteo Maximo", "Ciclo", "Año"]
    for col_num, header in enumerate(headers_calificaciones, 1):
        cell = ws.cell(row=start_row_calificaciones, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_aligned_text

    # Escribir la tabla de calificaciones
    for r_idx, row in enumerate(dataframe_to_rows(df_calificaciones, index=False, header=False), start_row_calificaciones + 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)

    # Ajustar el ancho de las columnas automáticamente
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = max_length + 2
        ws.column_dimensions[column_letter].width = adjusted_width

    # Guardar el libro en la respuesta
    wb.save(response)

    return response







def exportar_punteo_acumulado_excel(request, alumna_id):
    alumna = get_object_or_404(Alumna, id=alumna_id)

    # Obtener los filtros
    grado_id = request.GET.get('grado')
    year = request.GET.get('year')
    curso_id = request.GET.get('curso')

    # Imprimir valores para depuración
    print(f"Grado ID: {grado_id}")
    print(f"Año: {year}")
    print(f"Curso ID: {curso_id}")

    # Iniciar la consulta base
    calificaciones = CalificacionActividad.objects.filter(asignacion_ciclo__alumna=alumna)

    # Aplicar filtros
    if grado_id:
        calificaciones = calificaciones.filter(asignacion_ciclo__grado_id=grado_id)
    if year:
        calificaciones = calificaciones.filter(asignacion_ciclo__year=year)
    if curso_id:
        calificaciones = calificaciones.filter(actividad__curso_id=curso_id)

    # Imprimir el queryset para depuración
    print(calificaciones.query)

    # Calcular el punteo acumulado por curso, grado y año
    acumulado = calificaciones.values(
        'actividad__curso__nombre_curso',
        'asignacion_ciclo__grado__nombre_grado',
        'asignacion_ciclo__year'
    ).annotate(punteo_total=Sum('punteo')).order_by('actividad__curso__nombre_curso', 'asignacion_ciclo__year')

    # Convertir a DataFrame
    df_acumulado = pd.DataFrame(list(acumulado))

    # Verificar si el DataFrame está vacío
    if df_acumulado.empty:
        df_acumulado = pd.DataFrame(columns=['Curso', 'Punteo Total', 'Grado', 'Año'])
    else:
        # Renombrar columnas para que coincidan con los encabezados deseados
        df_acumulado = df_acumulado.rename(columns={
            'actividad__curso__nombre_curso': 'Curso',
            'asignacion_ciclo__grado__nombre_grado': 'Grado',
            'asignacion_ciclo__year': 'Año',
            'punteo_total': 'Punteo Total'
        })

    # Reordenar columnas
    df_acumulado = df_acumulado[['Curso', 'Punteo Total', 'Grado', 'Año']]

    # Imprimir el DataFrame para depuración
    print(df_acumulado)

    # Crear un archivo Excel con la información
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="punteo_acumulado_alumna.xlsx"'

    # Crear un libro y una hoja de trabajo
    wb = Workbook()
    ws = wb.active
    ws.title = 'Punteo Acumulado Alumna'

    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="008000", end_color="008000", fill_type="solid")
    center_aligned_text = Alignment(horizontal="center")
    data_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")

    # Escribir los encabezados de la alumna
    headers_alumna = ["Código", "Nombre", "Apellido"]
    for col_num, header in enumerate(headers_alumna, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_aligned_text

    # Escribir los datos de la alumna
    datos_alumna = {
        "Código": alumna.codigo,
        "Nombre": alumna.persona.nombre,
        "Apellido": alumna.persona.apellido
    }
    
    for col_num, header in enumerate(headers_alumna, 1):
        cell = ws.cell(row=2, column=col_num, value=datos_alumna[header])
        cell.fill = data_fill
        cell.alignment = center_aligned_text

    # Escribir una fila en blanco para separar los datos de la alumna de la tabla de punteo acumulado
    start_row_acumulado = 4

    # Escribir los encabezados de la tabla de punteo acumulado
    headers_acumulado = ['Curso', 'Punteo Total', 'Grado', 'Año']
    for col_num, header in enumerate(headers_acumulado, 1):
        cell = ws.cell(row=start_row_acumulado, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_aligned_text
        
    # Escribir la tabla de punteo acumulado
    for r_idx, row in df_acumulado.iterrows():
        for col_num, value in enumerate(row, 1):
            cell = ws.cell(row=start_row_acumulado + r_idx + 1, column=col_num, value=value)
            cell.alignment = center_aligned_text

    # Ajustar el ancho de las columnas automáticamente
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = max_length + 2
        ws.column_dimensions[column_letter].width = adjusted_width

    # Guardar el libro en la respuesta
    wb.save(response)

    return response



def generar_reporte_excel(grados_ids, year):
    # Crear nuevo libro de Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte de Notas"

    # Definir estilos
    header_style = Font(bold=True)
    center_align = Alignment(horizontal='center')
    header_fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")

    # Obtener grados y sus cursos
    grados = Grado.objects.filter(id__in=grados_ids, estado=True).prefetch_related('curso_set')
    
    if not grados:
        raise ValueError("No se encontraron grados para generar el reporte")

    # Obtener todos los cursos únicos de todos los grados
    cursos_unicos = set()
    for grado in grados:
        cursos = grado.curso_set.all()  # Considerar todos los cursos, sin importar el estado
        cursos_unicos.update(curso.nombre_curso for curso in cursos)
    cursos_unicos = sorted(list(cursos_unicos))  # Convertir a lista ordenada

    # Crear encabezados principales
    current_col = 2  # Primera columna es para nombres
    ws['A1'] = 'Alumna'
    ws['A1'].font = header_style
    
    # Crear encabezados por grado y curso
    for grado in grados:
        for curso in cursos_unicos:
            col_letter = get_column_letter(current_col)
            ws[f'{col_letter}1'] = grado.nombre_grado
            ws[f'{col_letter}2'] = curso
            ws[f'{col_letter}1'].font = header_style
            ws[f'{col_letter}2'].font = header_style
            ws[f'{col_letter}1'].alignment = center_align
            ws[f'{col_letter}2'].alignment = center_align
            ws[f'{col_letter}1'].fill = header_fill
            ws[f'{col_letter}2'].fill = header_fill
            current_col += 1

    # Agregar columnas de promedios por curso
    for curso in cursos_unicos:
        col_letter = get_column_letter(current_col)
        ws[f'{col_letter}1'] = f'Promedio {curso}'
        ws[f'{col_letter}1'].font = header_style
        ws[f'{col_letter}1'].alignment = center_align
        ws[f'{col_letter}1'].fill = header_fill
        current_col += 1

    # Columna final de estado
    col_letter = get_column_letter(current_col)
    ws[f'{col_letter}1'] = 'Estado'
    ws[f'{col_letter}1'].font = header_style
    ws[f'{col_letter}1'].alignment = center_align
    ws[f'{col_letter}1'].fill = header_fill

    # Obtener asignaciones y promociones
    asignaciones = AsignacionCiclo.objects.filter(
        grado_id__in=grados_ids, 
        year=year
    ).select_related(
        'alumna',
        'alumna__persona',
        'grado'
    )

    # Obtener todas las promociones para el año
    promociones = {
        (p.asignacion_ciclo.alumna_id, p.asignacion_ciclo.grado_id): p.aprobado 
        for p in Promocion.objects.filter(año=year)
    }

    # Agrupar asignaciones por alumna
    asignaciones_por_alumna = {}
    for asignacion in asignaciones:
        alumna_id = asignacion.alumna.id
        if alumna_id not in asignaciones_por_alumna:
            asignaciones_por_alumna[alumna_id] = {
                'nombre': f"{asignacion.alumna.persona.nombre} {asignacion.alumna.persona.apellido}",
                'asignaciones': {}
            }
        asignaciones_por_alumna[alumna_id]['asignaciones'][asignacion.grado.id] = asignacion

    # Procesar cada alumna
    current_row = 3  # Comenzar después de los encabezados
    for alumna_data in asignaciones_por_alumna.values():
        # Escribir nombre de la alumna
        ws[f'A{current_row}'] = alumna_data['nombre']
        
        col_index = 2
        promedios_cursos = {curso: [] for curso in cursos_unicos}
        
        # Procesar notas por grado y curso
        for grado in grados:
            asignacion = alumna_data['asignaciones'].get(grado.id)
            
            for curso_nombre in cursos_unicos:
                # Encontrar el curso correspondiente en el grado actual
                curso_actual = grado.curso_set.filter(
                    nombre_curso=curso_nombre
                ).first()
                
                if curso_actual:
                    # Obtener la suma de calificaciones sin filtrar por estado de actividad
                    calificaciones = CalificacionActividad.objects.filter(
                        asignacion_ciclo=asignacion,
                        actividad__curso=curso_actual
                    )

                    suma_calificaciones = calificaciones.aggregate(Sum('punteo'))['punteo__sum'] or 0
                    promedio = round(suma_calificaciones, 2)

                    # Guardar para el promedio final
                    promedios_cursos[curso_nombre].append(promedio)
                else:
                    promedio = 0
                
                # Escribir nota
                ws.cell(row=current_row, column=col_index, value=promedio)
                ws.cell(row=current_row, column=col_index).alignment = center_align
                col_index += 1

        # Calcular y escribir promedios por curso
        for curso_nombre, notas in promedios_cursos.items():
            if notas:
                promedio_final = sum(notas) / len(notas)
                ws.cell(row=current_row, column=col_index, value=round(promedio_final, 2))
            else:
                ws.cell(row=current_row, column=col_index, value=0)
            ws.cell(row=current_row, column=col_index).alignment = center_align
            col_index += 1

        # Obtener estado de la alumna usando la tabla de promociones
        estado = 'NO APROBADO'  # Valor predeterminado
        for grado in grados:
            if grado.id in alumna_data['asignaciones']:
                alumna_id = alumna_data['asignaciones'][grado.id].alumna.id
                grado_id = grado.id
                estado = 'APROBADO' if promociones.get((alumna_id, grado_id)) else 'NO APROBADO'
                break  # Tomar el primer grado donde existe una asignación

        ws.cell(row=current_row, column=col_index, value=estado)
        ws.cell(row=current_row, column=col_index).alignment = center_align

        current_row += 1

    # Ajustar ancho de columnas
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # Generar archivo
    from io import BytesIO
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    return excel_file
